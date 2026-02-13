from io import BytesIO
import re
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def _find_month_summary_sheet(wb):
    """
    워크북에서 '{m}월 총평' 시트를 찾아 (시트객체, m) 반환.
    없으면 (None, None).
    """
    for name in wb.sheetnames:
        m = re.match(r"^(\d{1,2})\s*월\s*총평$", str(name).strip())
        if m:
            month = int(m.group(1))
            return wb[name], month
    return None, None


def _get_month_col_in_by_tier(month: int, year: int) -> int:
    """
    by Tier 시트의 월 컬럼 계산.
    템플릿 기준:
      - Jan-25 시작 열 = O(15)
      - Jan-26 시작 열 = AA(27)
      - Jan-27 시작 열 = AM(39)
    """
    if not (1 <= month <= 12):
        raise ValueError(f"유효하지 않은 월입니다: {month}")
    if year < 2025:
        raise ValueError(f"지원하지 않는 연도입니다: {year}")

    start_col_2025 = 15  # O
    start_col = start_col_2025 + (year - 2025) * 12
    return start_col + (month - 1)


def _num(v):
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", ""))
    except Exception:
        return 0


# -----------------------------
# 총평 시트 수식 평가(부분 지원)
# - openpyxl은 수식을 계산하지 않으므로, 총평 셀의 값이 None/0 캐시일 때
#   아래 제한된 수식 패턴(셀 참조, COUNTIF, SUM, COUNTA, IFERROR(VLOOKUP))을 파이썬으로 평가한다.
# -----------------------------
_REF_RE = re.compile(r"^\s*=?\s*(?:'([^']+)'|([^'!]+))!\s*([A-Z]{1,3}\d+)\s*$")
_COUNTIF_RE = re.compile(
    r'^\s*=?\s*COUNTIF\(\s*'
    r"(?:'([^']+)'|([^'!]+))!\s*([A-Z]{1,3})(\d+)\s*:\s*([A-Z]{1,3})(\d+|1048576)\s*,\s*\"([^\"]*)\"\s*\)\s*$",
    re.I
)
_SUM_RE = re.compile(
    r'^\s*=?\s*SUM\(\s*'
    r'\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)\s*\)\s*$',
    re.I
)
_COUNTA_RE = re.compile(
    r'^\s*=?\s*COUNTA\(\s*'
    r'\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)\s*\)\s*$',
    re.I
)

# IFERROR(VLOOKUP(D7,'Tier Table'!$B:$C,2,FALSE),0)
_VLOOKUP_IFERROR_RE = re.compile(
    r'^\s*IFERROR\(\s*VLOOKUP\(\s*([A-Z]{1,3}\d+)\s*,\s*'
    r"(?:'([^']+)'|([^'!]+))!\$?([A-Z]{1,3})\s*:\s*\$?([A-Z]{1,3})\s*,\s*2\s*,\s*FALSE\s*\)\s*,\s*0\s*\)\s*$",
    re.I
)

# Tier Table 맵 캐시 (워크북 객체별)
_TIER_MAP_CACHE = {}


def _ws_get(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트가 없습니다: {sheet_name}")
    return wb[sheet_name]


def _countif_range(wb_vals, sheet_name: str, col_letter: str, start_row: int, end_row: int, criterion: str):
    ws = _ws_get(wb_vals, sheet_name)

    # end_row가 1048576 같은 최대치일 수 있으니 실제 행수/합리 범위로 제한
    max_scan = min(ws.max_row or start_row, end_row if end_row < 10**7 else (ws.max_row or start_row))
    max_scan = min(max_scan, start_row + 50000)  # 안전 상한

    count = 0
    empty_streak = 0

    for r in range(start_row, max_scan + 1):
        v = ws[f"{col_letter}{r}"].value

        # 데이터 끝 감지: D/E가 둘 다 빈 행이 연속되면 종료
        d = ws[f"D{r}"].value
        e = ws[f"E{r}"].value
        if (d is None or str(d).strip() == "") and (e is None or str(e).strip() == ""):
            empty_streak += 1
            if empty_streak >= 20:
                break
            continue
        empty_streak = 0

        if v is not None and str(v).strip() == criterion:
            count += 1

    return count


def _eval_countif_formula(wb_vals, formula: str):
    # 예: COUNTIF(CP_1_work!D5:D1048576,"연합뉴스")
    fm = formula.strip().lstrip("=")
    m = _COUNTIF_RE.match(fm)
    if not m:
        return None

    sheet_quoted = m.group(1)
    sheet_plain = m.group(2)
    sheet_name = sheet_quoted or sheet_plain

    col1 = m.group(3)
    r1 = int(m.group(4))
    col2 = m.group(5)
    r2 = int(m.group(6)) if m.group(6).isdigit() else 1048576
    criterion = m.group(7)

    if col1 != col2:
        return None

    return _countif_range(wb_vals, sheet_name, col1, r1, r2, criterion)


def _build_tier_maps(wb_vals):
    """
    Tier Table 시트에서
      - Tier1: B -> C
      - Tier2: D -> E
    맵을 만든다.
    """
    if "Tier Table" not in wb_vals.sheetnames:
        raise ValueError("검수파일에 'Tier Table' 시트가 없습니다.")

    ws = wb_vals["Tier Table"]

    tier1 = {}
    tier2 = {}

    # 보통 3행부터 값이 시작하니 넉넉히 스캔
    for r in range(3, (ws.max_row or 3) + 1):
        k1 = ws[f"B{r}"].value
        v1 = ws[f"C{r}"].value
        if k1 is not None and str(k1).strip() != "":
            tier1[str(k1).strip()] = _num(v1)

        k2 = ws[f"D{r}"].value
        v2 = ws[f"E{r}"].value
        if k2 is not None and str(k2).strip() != "":
            tier2[str(k2).strip()] = _num(v2)

    return tier1, tier2


def _eval_iferror_vlookup(wb_vals, wb_formulas, sheet_name: str, formula: str, tier1_map, tier2_map):
    """
    IFERROR(VLOOKUP(D7,'Tier Table'!$B:$C,2,FALSE),0) 형태를 0/1(혹은 테이블 값)로 평가
    """
    fm = formula.strip().lstrip("=")
    m = _VLOOKUP_IFERROR_RE.match(fm)
    if not m:
        return None

    key_addr = m.group(1)  # 예: D7
    table_sheet = m.group(2) or m.group(3)  # 'Tier Table'
    col_left = m.group(4)   # B or D
    col_right = m.group(5)  # C or E

    if table_sheet != "Tier Table":
        return None

    ws = _ws_get(wb_vals, sheet_name)
    key = ws[key_addr].value
    key = (str(key).strip() if key is not None else "")
    if key == "":
        return 0

    if col_left.upper() == "B" and col_right.upper() == "C":
        return tier1_map.get(key, 0)
    if col_left.upper() == "D" and col_right.upper() == "E":
        return tier2_map.get(key, 0)

    return None


def _get_cell_numeric_with_fallback(wb_vals, wb_formulas, sheet_name: str, addr: str) -> float:
    """
    data_only=True에서 값이 비거나(또는 캐시가 0이라도) 수식이 있으면,
    data_only=False의 수식을 파싱해 참조/부분평가로 숫자를 만든다.

    지원:
      - =Sheet!A1 (단순 참조, 중첩 참조 재귀)
      - =COUNTIF(Sheet!D5:D1048576,"연합뉴스")
      - =SUM(...), =COUNTA(...)
      - =IFERROR(VLOOKUP(...),0)  (Tier Table)
    """
    ws_v = _ws_get(wb_vals, sheet_name)
    ws_f = _ws_get(wb_formulas, sheet_name)

    v_cached = ws_v[addr].value
    f = ws_f[addr].value

    # ✅ 수식이 없는 셀만 캐시값을 그대로 사용
    if not (isinstance(f, str) and f.strip().startswith("=")):
        if v_cached is None or str(v_cached).strip() == "":
            return 0
        return _num(v_cached)

    # ✅ Tier Table 맵(워크북별 캐시)
    cache_key = id(wb_vals)
    if cache_key not in _TIER_MAP_CACHE:
        _TIER_MAP_CACHE[cache_key] = _build_tier_maps(wb_vals)
    tier1_map, tier2_map = _TIER_MAP_CACHE[cache_key]

    # ✅ 수식이 있는 셀은 캐시가 0이어도 믿지 말고 수식을 따라간다
    formula = f.strip()

    # 1) 단순 참조 (=CP_1_work!F3 같은 케이스) -> 재귀로 따라가기
    mref = _REF_RE.match(formula.lstrip("="))
    if mref:
        ref_sheet = mref.group(1) or mref.group(2)
        ref_addr = mref.group(3)
        return _get_cell_numeric_with_fallback(wb_vals, wb_formulas, ref_sheet, ref_addr)

    # 2) COUNTIF
    cnt = _eval_countif_formula(wb_vals, formula)
    if cnt is not None:
        return _num(cnt)

    # 2.5) IFERROR(VLOOKUP(...),0) (Tier Table)
    vlu = _eval_iferror_vlookup(wb_vals, wb_formulas, sheet_name, formula, tier1_map, tier2_map)
    if vlu is not None:
        return _num(vlu)

    # 3) SUM/COUNTA: Sheet! prefix 분리 지원
    f_no_eq = formula.lstrip("=")

    m_pref = re.match(r"^(SUM|COUNTA)\(\s*(?:'([^']+)'|([^'!]+))!(.+)\)\s*$", f_no_eq, re.I)
    if m_pref:
        func = m_pref.group(1).upper()
        sheet_prefix = m_pref.group(2) or m_pref.group(3)
        range_part = m_pref.group(4).strip()
        inner = f"{func}({range_part})"
        tgt_sheet = sheet_prefix
    else:
        inner = f_no_eq
        tgt_sheet = sheet_name

    # SUM
    msum = _SUM_RE.match(inner)
    if msum:
        c1, r1, c2, r2 = msum.group(1), int(msum.group(2)), msum.group(3), int(msum.group(4))
        ws = _ws_get(wb_vals, tgt_sheet)
        min_col = column_index_from_string(c1)
        max_col = column_index_from_string(c2)
        total = 0

        for rr in range(r1, r2 + 1):
            for cc in range(min_col, max_col + 1):
                cell_addr = f"{get_column_letter(cc)}{rr}"
                # ✅ SUM 범위 내 셀도 수식일 수 있으니 fallback로 평가
                total += _get_cell_numeric_with_fallback(wb_vals, wb_formulas, tgt_sheet, cell_addr)

        return total

    # COUNTA
    mca = _COUNTA_RE.match(inner)
    if mca:
        c1, r1, c2, r2 = mca.group(1), int(mca.group(2)), mca.group(3), int(mca.group(4))
        ws = _ws_get(wb_vals, tgt_sheet)
        min_col = column_index_from_string(c1)
        max_col = column_index_from_string(c2)
        cnt2 = 0
        for rr in range(r1, r2 + 1):
            for cc in range(min_col, max_col + 1):
                vv_addr = f"{get_column_letter(cc)}{rr}"
                vv = _get_cell_numeric_with_fallback(wb_vals, wb_formulas, tgt_sheet, vv_addr)
                # COUNTA는 "비어있지 않은 셀" 기준인데,
                # 여기서는 숫자 평가 기반이므로 0도 값으로 볼지 애매함.
                # 템플릿에서 COUNTA를 쓰는 경우는 대개 문자열 카운트라 본 로직은 보수적으로 유지.
                if vv != 0:
                    cnt2 += 1
        return cnt2

    # 여기까지 왔다는 건 수식 형태가 지원 밖
    if v_cached is not None and str(v_cached).strip() != "":
        return _num(v_cached)

    raise ValueError(f"[총평 시트] {addr} 수식을 해석할 수 없습니다: {formula}")


# -----------------------------
# by Coverage 업데이트용 유틸
# -----------------------------
def _make_month_label(year: int, month: int) -> str:
    # 예: 2025, 12 -> "Dec-25"
    if not (1 <= month <= 12):
        raise ValueError("month must be 1~12")
    yy = str(year)[-2:]
    return f"{MONTH_ABBR[month-1]}-{yy}"


def _cell_to_month_label(v):
    """
    셀 값이
    - datetime/date -> 'Dec-25'
    - 'Dec-25' 문자열 -> 'Dec-25'
    그 외 -> None
    """
    if isinstance(v, (datetime, date)):
        return f"{MONTH_ABBR[v.month - 1]}-{str(v.year)[-2:]}"
    if isinstance(v, str):
        return v.strip()
    return None


def _find_row_by_label(ws, label: str, label_col: int = 1):
    for r in range(1, ws.max_row + 1):
        raw = ws.cell(row=r, column=label_col).value
        cell_label = _cell_to_month_label(raw)
        if cell_label == label:
            return r
    return None


def _contains_any(text: str, keywords) -> bool:
    t = (text or "").strip().lower()
    return any(k.lower() in t for k in keywords)


def _calc_coverage_from_work_sheet(checked_wb, *, sheet_name: str, tv_keywords):
    """
    지정한 *_work 시트에서 M7:N50을 읽어 6개 카테고리 합계를 반환:
    [Smartphone, AI, TV/Display, Semiconductor, Auto, IoT]
    """
    if sheet_name not in checked_wb.sheetnames:
        raise ValueError(f"검수완료파일에 '{sheet_name}' 시트가 없습니다.")

    ws = checked_wb[sheet_name]

    sums = {
        "smartphone": 0,
        "ai": 0,
        "tv_display": 0,
        "semi": 0,
        "auto": 0,
        "iot": 0,
    }

    for r in range(7, 51):  # 7~50
        m_val = _num(ws[f"M{r}"].value)
        n_txt = ws[f"N{r}"].value
        s = (str(n_txt) if n_txt is not None else "").strip()

        if "스마트폰" in s:
            sums["smartphone"] += m_val
        if "ai" in s.lower():
            sums["ai"] += m_val
        if _contains_any(s, tv_keywords):
            sums["tv_display"] += m_val
        if "반도체" in s:
            sums["semi"] += m_val
        if "전기차" in s:
            sums["auto"] += m_val
        if "iot" in s.lower():
            sums["iot"] += m_val

    return [
        sums["smartphone"],
        sums["ai"],
        sums["tv_display"],
        sums["semi"],
        sums["auto"],
        sums["iot"],
    ]


def _calc_omdia_tv_from_cp_work(checked_wb, *, month: int, keywords):
    """
    CP_{m}_work 시트에서 M7:N50 중, N열 텍스트에 keywords가 포함된 행들의 M열 합계를 반환.
    (Omdia TV 단일 값)
    """
    sheet_name = f"CP_{month}_work"
    if sheet_name not in checked_wb.sheetnames:
        raise ValueError(f"검수완료파일에 '{sheet_name}' 시트가 없습니다.")

    ws = checked_wb[sheet_name]
    total = 0

    for r in range(7, 51):
        m_val = _num(ws[f"M{r}"].value)
        n_txt = ws[f"N{r}"].value
        s = (str(n_txt) if n_txt is not None else "").strip()
        if _contains_any(s, keywords):
            total += m_val

    return total


def _write_coverage_block_to_master(
    master_wb,
    *,
    year: int,
    month: int,
    start_col: int,   # B=2, H=8, N=14 ...
    values,
    sheet_name: str = "by Coverage",
):
    """
    Master의 'by Coverage' 시트에서 해당 월 라벨 행을 찾아 start_col부터 values를 순서대로 입력.
    """
    if sheet_name not in master_wb.sheetnames:
        raise ValueError(f"Master 파일에 '{sheet_name}' 시트가 없습니다.")

    ws = master_wb[sheet_name]
    label = _make_month_label(year, month)  # 예: Jan-26
    row = _find_row_by_label(ws, label, label_col=1)

    if row is None:
        raise ValueError(f"Master by Coverage에서 '{label}' 행을 찾을 수 없습니다.")

    for i, v in enumerate(values):
        ws.cell(row=row, column=start_col + i, value=v)


# -----------------------------
# by Tier 업데이트
# -----------------------------
def _read_summary_EFG(wb_vals, wb_formulas, summary_sheet_name: str, row: int):
    """
    총평 시트 row에서 E/F/G를 읽되, 값이 비면(수식 캐시 없음) 수식을 부분 평가해 숫자로 반환.
    템플릿 의미:
      - E = 연합뉴스 수
      - F = Tier1(연합 제외) 수
      - G = Tier2 수
    """
    e = _get_cell_numeric_with_fallback(wb_vals, wb_formulas, summary_sheet_name, f"E{row}")
    f = _get_cell_numeric_with_fallback(wb_vals, wb_formulas, summary_sheet_name, f"F{row}")
    g = _get_cell_numeric_with_fallback(wb_vals, wb_formulas, summary_sheet_name, f"G{row}")
    return e, f, g


def _write_by_tier_block(
    tier_ws,
    *,
    month_col: int,
    src_e_yonhap: float,
    src_f_tier1_excl: float,
    src_g_tier2: float,
    dst_rows: tuple
):
    """
    dst_rows = (row_yonhap, row_tier1_excl, row_tier2, row_total)

    ✅ 업데이트 규칙:
      - 연합뉴스: 총평 E 그대로
      - Tier1(연합 제외): (총평 F) - (총평 E)
      - Tier2: 총평 G 그대로
      - 합계: E + (F-E) + G
    """
    r_y, r_t1, r_t2, r_total = dst_rows

    tier1_excl = _num(src_f_tier1_excl) - _num(src_e_yonhap)

    # (방어) 음수면 0으로 클램프 (원하면 제거 가능)
    if tier1_excl < 0:
        tier1_excl = 0

    tier_ws.cell(row=r_y,  column=month_col, value=_num(src_e_yonhap))
    tier_ws.cell(row=r_t1, column=month_col, value=tier1_excl)
    tier_ws.cell(row=r_t2, column=month_col, value=_num(src_g_tier2))

    # 합계: E + (F-E) + G
    col_letter = get_column_letter(month_col)
    tier_ws.cell(row=r_total, column=month_col, value=f"=SUM({col_letter}{r_y}:{col_letter}{r_t2})")


def process_master_update(checked_bytes: bytes, master_bytes: bytes, *, year: int, month: int) -> bytes:
    """
    Step3:
    1) 검수완료파일 '{m}월 총평'에서 by Tier 업데이트 (총평 시트 기반)
       - 총평/워크시트 수식 캐시가 없을 경우를 대비해, 제한 수식 패턴을 파이썬으로 부분 평가.
       - 트렌드포스 포함 매핑:
         총평 row: CP(5), 트렌드포스(6), IDC(7), OmdiaTV(8), DSCC(9)
         by Tier row block:
           CP        : 3~6
           트렌드포스: 8~11
           IDC       : 13~16
           OmdiaTV   : 18~21
           DSCC      : 23~26

       매핑:
         E -> (3,8,13,18,23)
         F -> (4,9,14,19,24)
         G -> (5,10,15,20,25)
         SUM -> (6,11,16,21,26)

    2) by Coverage 업데이트(기존 로직 유지)
       - Counterpoint: CP_{m}_work -> B~G (6개)
       - IDC        : IDC_{m}_work -> H~M (6개)
       - Omdia TV   : CP_{m}_work -> N (1개)
    """

    # 1) 검수완료파일 로드
    checked_wb_vals = load_workbook(BytesIO(checked_bytes), data_only=True)
    checked_wb_formulas = load_workbook(BytesIO(checked_bytes), data_only=False)

    summary_ws_f, found_month = _find_month_summary_sheet(checked_wb_formulas)
    if summary_ws_f is None or found_month is None:
        raise ValueError("검수완료파일에서 '{m}월 총평' 시트를 찾을 수 없습니다.")

    summary_sheet_name = summary_ws_f.title
    if summary_sheet_name not in checked_wb_vals.sheetnames:
        raise ValueError(f"검수완료파일에서 총평 시트('{summary_sheet_name}')를 찾을 수 없습니다.")

    if found_month != month:
        raise ValueError(f"기간의 월({month})과 검수파일 총평 시트의 월({found_month})이 일치하지 않습니다.")

    # 2) 마스터 파일 로드
    master_wb = load_workbook(BytesIO(master_bytes), data_only=False)

    # 3) by Tier 업데이트
    if "by Tier" not in master_wb.sheetnames:
        raise ValueError("Master 파일에 'by Tier' 시트가 없습니다.")
    tier_ws = master_wb["by Tier"]

    month_col = _get_month_col_in_by_tier(month, year=year)

    mappings = [
        ("CP",         5, (3, 4, 5, 6)),
        ("트렌드포스",  6, (8, 9, 10, 11)),
        ("IDC",        7, (13, 14, 15, 16)),
        ("OmdiaTV",    8, (18, 19, 20, 21)),
        ("DSCC",       9, (23, 24, 25, 26)),
    ]

    for _vendor_label, summary_row, dst_rows in mappings:
        e, f, g = _read_summary_EFG(checked_wb_vals, checked_wb_formulas, summary_sheet_name, summary_row)
        _write_by_tier_block(
            tier_ws,
            month_col=month_col,
            src_e_yonhap=e,
            src_f_tier1_excl=f,
            src_g_tier2=g,
            dst_rows=dst_rows,
        )

    # 4) by Coverage 업데이트
    # Counterpoint (B~G)
    cp_tv_keywords = ["tv", "디스플레이", "lcd", "led", "모니터", "oled", "xr"]
    cp_values6 = _calc_coverage_from_work_sheet(
        checked_wb_vals,
        sheet_name=f"CP_{month}_work",
        tv_keywords=cp_tv_keywords,
    )
    _write_coverage_block_to_master(
        master_wb,
        year=year,
        month=month,
        start_col=2,   # B
        values=cp_values6,
    )

    # IDC (H~M)
    idc_tv_keywords = ["tv", "디스플레이", "lcd", "led", "모니터", "oled", "xr"]
    idc_values6 = _calc_coverage_from_work_sheet(
        checked_wb_vals,
        sheet_name=f"IDC_{month}_work",
        tv_keywords=idc_tv_keywords,
    )
    _write_coverage_block_to_master(
        master_wb,
        year=year,
        month=month,
        start_col=8,   # H
        values=idc_values6,
    )

    # Omdia TV (N 단일 값)
    omdia_keywords = ["tv", "디스플레이", "oled", "lcd"]
    omdia_tv_val = _calc_omdia_tv_from_cp_work(
        checked_wb_vals,
        month=month,
        keywords=omdia_keywords,
    )
    _write_coverage_block_to_master(
        master_wb,
        year=year,
        month=month,
        start_col=14,  # N
        values=[omdia_tv_val],
    )

    # 5) 저장 후 반환
    out = BytesIO()
    master_wb.save(out)
    out.seek(0)
    return out.getvalue()
