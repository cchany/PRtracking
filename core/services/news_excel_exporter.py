# core/services/news_excel_exporter.py
from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from io import BytesIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .naver_news_client import NewsItem
from .news_classifier import Classified


HEADER = [
    "company",
    "pub_date",
    "press",
    "category",
    "score",
    "matched_keywords",
    "title",
    "description",
    "originallink",
    "naver_link",
]


def _auto_width(ws, max_col: int, max_rows: int = 2000):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 10
        for r in range(1, min(ws.max_row, max_rows) + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[letter].width = min(max_len + 2, 55)


def build_news_workbook(rows: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()

    # Sheet 1: news_all
    ws = wb.active
    ws.title = "news_all"

    bold = Font(bold=True)
    center = Alignment(vertical="center", wrap_text=True)

    for i, h in enumerate(HEADER, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold

    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=row.get("company"))
        # pub_date -> YYYY-MM-DD HH:MM
        pd: datetime = row.get("pub_date")
        ws.cell(row=r_idx, column=2, value=pd.strftime("%Y-%m-%d %H:%M") if isinstance(pd, datetime) else "")
        ws.cell(row=r_idx, column=3, value=row.get("press"))
        ws.cell(row=r_idx, column=4, value=row.get("category"))
        ws.cell(row=r_idx, column=5, value=row.get("score"))
        ws.cell(row=r_idx, column=6, value=", ".join(row.get("matched_keywords") or []))
        ws.cell(row=r_idx, column=7, value=row.get("title"))
        ws.cell(row=r_idx, column=8, value=row.get("description"))

        ol = row.get("originallink") or ""
        nl = row.get("naver_link") or ""
        ws.cell(row=r_idx, column=9, value=ol)
        ws.cell(row=r_idx, column=10, value=nl)

    ws.freeze_panes = "A2"
    for col in range(1, len(HEADER) + 1):
        ws.cell(row=1, column=col).alignment = center

    # Sheet 2: summary
    ws2 = wb.create_sheet("summary")
    ws2["A1"] = "company"
    ws2["B1"] = "category"
    ws2["C1"] = "count"
    ws2["A1"].font = ws2["B1"].font = ws2["C1"].font = bold

    counter = {}
    for row in rows:
        key = (row.get("company"), row.get("category"))
        counter[key] = counter.get(key, 0) + 1

    i = 2
    for (company, category), cnt in sorted(counter.items(), key=lambda x: (x[0][0], x[0][1])):
        ws2.cell(row=i, column=1, value=company)
        ws2.cell(row=i, column=2, value=category)
        ws2.cell(row=i, column=3, value=cnt)
        i += 1

    ws2.freeze_panes = "A2"
    _auto_width(ws, len(HEADER))
    _auto_width(ws2, 3)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
