from io import BytesIO
from collections import Counter
from openpyxl import load_workbook
from openpyxl.chart import PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
import re
from datetime import datetime, timezone, timedelta
from openpyxl.styles import Font

BASE_SOURCES = ["CP", "트렌드포스", "IDC", "OmdiaTV", "DSCC"]


def _find_main_and_work_pairs(wb):
    """
    워크북에서 각 소스별 (원본 시트, _work 시트) 쌍을 찾아 반환.

    예:
      CP_10, CP_10_work
      IDC_10, IDC_10_work ...
    """
    pairs = []
    for base in BASE_SOURCES:
        main_sheet = None
        for name in wb.sheetnames:
            if name.startswith(base + "_") and not name.endswith("_work"):
                main_sheet = name
                break
        if not main_sheet:
            continue
        work_sheet = f"{main_sheet}_work"
        if work_sheet in wb.sheetnames:
            pairs.append((main_sheet, work_sheet))
    return pairs


def _copy_main_to_work(main_ws, work_ws):
    """
    1. CP_{m}, IDC_{m}, OmdiaTV_{m}, DSCC_{m} 각 시트의 B5:E2000을 복사
    2. CP_{m}_work, ... 각 시트의 C7:F1002에 붙여넣기
    """
    src_row_start, src_row_end = 5, 2000
    src_col_start = 2  # B
    dst_row_start = 7
    dst_col_start = 3  # C

    for r in range(src_row_start, src_row_end + 1):
        dst_r = dst_row_start + (r - src_row_start)
        for offset in range(4):  # B~E → C~F
            src_c = src_col_start + offset
            dst_c = dst_col_start + offset
            val = main_ws.cell(row=r, column=src_c).value
            work_ws.cell(row=dst_r, column=dst_c, value=val)


def _find_tier_table_sheet(wb):
    """
    Tier Table 시트 찾기.
    기본적으로 'Tier Table'을 우선으로 하고,
    없으면 'tier'가 들어간 첫 시트를 사용.
    """
    if "Tier Table" in wb.sheetnames:
        return wb["Tier Table"]

    for name in wb.sheetnames:
        if "tier" in name.replace(" ", "").lower():
            return wb[name]
    return None


def _update_tier_table_if_needed(wb, work_ws):
    """
    D3와 F2의 숫자가 다르면,
    G/H가 둘 다 0인 행의 '언론사(D열)'를 Tier Table 시트의 D열 맨 아래에 추가.
    """

    def _to_int(v):
        if v is None:
            return None
        try:
            return int(str(v).replace(",", ""))
        except Exception:
            return None

    d3_val = _to_int(work_ws["D3"].value)
    f2_val = _to_int(work_ws["F2"].value)

    # 숫자가 같으면 Tier Table 업데이트 불필요
    if d3_val is not None and f2_val is not None and d3_val == f2_val:
        return

    tier_ws = _find_tier_table_sheet(wb)
    if tier_ws is None:
        return

    # Tier Table 내 기존 언론사( Tier1: B열, Tier2: D열 ) 수집
    existing_names = set()
    for row in range(2, 5000):
        for col in (2, 4):  # B, D
            v = tier_ws.cell(row=row, column=col).value
            if v is None:
                continue
            s = str(v).strip()
            if s:
                existing_names.add(s)

    def _to_float_zero(v):
        if v is None or str(v).strip() == "":
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(str(v).replace(",", ""))
        except Exception:
            return 0.0

    # 신규 언론사 후보: G/H 둘 다 0이고, 언론사(D열)가 기존 목록에 없는 것
    candidates = []
    for r in range(7, 1003):
        t1 = _to_float_zero(work_ws.cell(row=r, column=7).value)  # G열(Tier1)
        t2 = _to_float_zero(work_ws.cell(row=r, column=8).value)  # H열(Tier2)
        if t1 != 0.0 or t2 != 0.0:
            continue

        press = work_ws.cell(row=r, column=4).value  # 🔴 언론사: D열
        if not press:
            continue
        press_str = str(press).strip()
        if press_str and press_str not in existing_names and press_str not in candidates:
            candidates.append(press_str)

    if not candidates:
        return

    # Tier Table 시트 D열에서 첫 번째 빈 행 찾기
    row = 2
    while tier_ws.cell(row=row, column=4).value not in (None, ""):
        row += 1

    # D열에 순서대로 추가
    for name in candidates:
        tier_ws.cell(row=row, column=4, value=name)
        row += 1


def _fill_categories_and_counts(main_ws, work_ws):
    """
    4. 원본 시트의 카테고리(G5:G800)를 읽어
       - 중복 제거된 카테고리 리스트를 L7:L...에 채우고
       - 각 카테고리 건수를 K열에 적음
    """
    # 기존 카테고리/건수 영역 초기화 (잔여 값 제거)
    for r in range(7, 2000):
        work_ws.cell(row=r, column=11, value=None)  # K
        work_ws.cell(row=r, column=12, value=None)  # L

    # 원본 시트에서 카테고리 전체 목록 + 빈값 제외
    categories = []
    for r in range(5, 2000):
        val = main_ws.cell(row=r, column=7).value  # G
        if val is None or str(val).strip() == "":
            continue
        categories.append(str(val).strip())

    freq = Counter(categories)

    # 중복 제거된 카테고리(등장 순서 유지)
    seen = set()
    unique_cats = []
    for cat in categories:
        if cat not in seen:
            seen.add(cat)
            unique_cats.append(cat)

    # L7부터 카테고리, K7부터 건수 입력
    row = 7
    for cat in unique_cats:
        work_ws.cell(row=row, column=12, value=cat)           # L
        work_ws.cell(row=row, column=11, value=freq.get(cat)) # K
        row += 1


def _sort_counts_to_MN(work_ws):
    """
    5. K7:L800 → 건수/카테고리 데이터를 읽어
       M7:N800에 '건수 기준 내림차순'으로 재정렬된 결과를 채운다.
    """
    # 기존 정렬 영역 초기화 (잔여 값 제거)
    for r in range(7, 2000):
        work_ws.cell(row=r, column=13, value=None)  # M
        work_ws.cell(row=r, column=14, value=None)  # N

    rows = []
    for r in range(7, 2000):
        count = work_ws.cell(row=r, column=11).value  # K
        cat = work_ws.cell(row=r, column=12).value    # L
        if cat is None or str(cat).strip() == "":
            continue
        try:
            c_val = float(count) if count is not None else 0.0
        except Exception:
            c_val = 0.0
        rows.append((c_val, cat))

    if not rows:
        return

    # 건수 기준 내림차순 정렬
    rows.sort(key=lambda x: x[0], reverse=True)

    # M7:N... 에 채우기 (M: 건수, N: 카테고리)
    r_idx = 7
    for c_val, cat in rows:
        work_ws.cell(row=r_idx, column=13, value=c_val)  # M
        work_ws.cell(row=r_idx, column=14, value=cat)    # N
        r_idx += 1


def _prepare_chart_area(work_ws):
    """
    6. 원 그래프용 상위 카테고리 + 기타 영역 생성 + 차트 설정

    - N7:N14 → P7:P14 (카테고리)
    - M7:M14 → Q7:Q14 (건수)
    - M15:M30의 건수를 합산해 Q15에 넣고, P15에는 '기타' 입력
    - Q7:Q15의 표시형식을 0"건" 으로 지정
    - 첫 번째 PieChart가 있으면:
        · 데이터 레이블: 값 + 지시선 표시
        · 레이블 텍스트는 셀 값(= "n건") 사용
        · 색상 팔레트/스타일은 템플릿에서 지정한 그대로 유지
    """
    # 기존 차트 데이터 영역 초기화 (잔여 값/서식 제거)
    for r in range(7, 16):
        work_ws.cell(row=r, column=16, value=None)  # P
        q_cell = work_ws.cell(row=r, column=17, value=None)  # Q
        q_cell.number_format = "General"

    # 1) 상위 8개 복사 (M/N -> P/Q)
    for r in range(7, 15):
        cat = work_ws.cell(row=r, column=14).value  # N 열: 카테고리
        cnt = work_ws.cell(row=r, column=13).value  # M 열: 건수
        work_ws.cell(row=r, column=16, value=cat)   # P 열
        work_ws.cell(row=r, column=17, value=cnt)   # Q 열

    # 2) 기타 합산 (M15:M30)
    etc_sum = 0.0
    for r in range(15, 31):
        val = work_ws.cell(row=r, column=13).value  # M 열
        if isinstance(val, (int, float)):
            etc_sum += float(val)
        elif isinstance(val, str):
            try:
                etc_sum += float(val.replace(",", ""))
            except Exception:
                continue

    if etc_sum > 0:
        work_ws["P15"] = "기타"
        work_ws["Q15"] = etc_sum
    else:
        work_ws["P15"] = None
        work_ws["Q15"] = None

    # 3) Q7:Q15 숫자 서식 "0\"건\"" 적용 → 데이터 레이블이 "148건"처럼 보이도록
    for r in range(7, 16):
        cell = work_ws.cell(row=r, column=17)  # Q 열
        if cell.value not in (None, ""):
            cell.number_format = '0"건"'

    # 4) 첫 번째 원그래프가 있으면 데이터/라벨 옵션 재설정
    charts = getattr(work_ws, "_charts", [])
    if not charts:
        return

    data_ref = Reference(work_ws, min_col=17, max_col=17, min_row=7, max_row=15)  # Q7:Q15
    cat_ref = Reference(work_ws, min_col=16, max_col=16, min_row=7, max_row=15)   # P7:P15

    for ch in charts:
        if not isinstance(ch, PieChart):
            continue

        # 기존 시리즈 비우고 새 데이터로 연결 (차트 스타일/색상은 그대로 둠)
        ch.series = []
        ch.add_data(data_ref, titles_from_data=False)
        ch.set_categories(cat_ref)

        # 데이터 레이블 옵션: 값 + 지시선 표시
        dl = DataLabelList()
        dl.showVal = True
        dl.showLeaderLines = True
        dl.showPercent = False
        dl.showLegendKey = False
        dl.showCatName = False
        dl.showSerName = False
        dl.showBubbleSize = False
        dl.showRange = False
        dl.showLabel = False
        ch.dataLabels = dl

        break  # 첫 번째 PieChart만 처리


def _seoul_now_year_month():
    # Asia/Seoul = UTC+9 (고정 오프셋로 처리)
    now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=9)))
    return now.year, now.month


def _calc_year_month_row(year: int, month: int, base_year: int = 2020, base_row: int = 29) -> int:
    """
    C29=2020년 1월, C30=2020년 2월 ... 규칙 기반 행 계산
    row = base_row + (year-base_year)*12 + (month-1)
    """
    return base_row + (year - base_year) * 12 + (month - 1)


def _seoul_now_year_month():
    # Asia/Seoul = UTC+9 (고정 오프셋)
    now = datetime.now(timezone.utc).astimezone(timezone(timedelta(hours=9)))
    return now.year, now.month


def _calc_year_month_row(year: int, month: int, base_year: int = 2020, base_row: int = 29) -> int:
    """
    C29=2020년 1월, C30=2020년 2월 ... 규칙 기반 행 계산
    row = base_row + (year-base_year)*12 + (month-1)
    """
    return base_row + (year - base_year) * 12 + (month - 1)


def _update_month_summary_sheet(wb):
    """
    '{m}월 총평' 시트가 있으면,
    1) D5:D8, E5:E8, F5:F8, G5:G8 수식을 채워 넣고
    2) Chart1 데이터 범위를 강제 지정한다.
    3) B24:G136 테이블에서 '이번 달(현재 년/월)' 행을 찾아
       D~G에 (CP/IDC/OmdiaTV/DSCC) 총 게재 수를 자동 기록한다. (=D5~D8 링크)
    4) 이번달 vs 전월 게재 수 차이를 J24:M24에 기록(빨강+bold)한다.
    5) Chart2 데이터 범위를 '작년 1월 ~ 이번달'로 강제 지정한다.
       - categories: C(작년1월) ~ C(이번달)
       - series: D~G (계열명은 D24:G24 고정)
    """

    summary_ws = None
    month_num = None

    # 1) '{m}월 총평' 시트 찾기
    for name in wb.sheetnames:
        mm = re.match(r"^(\d{1,2})\s*월\s*총평$", str(name).strip())
        if mm:
            summary_ws = wb[name]
            month_num = int(mm.group(1))
            break

    if summary_ws is None or month_num is None:
        return

    m = month_num

    sources = [
        ("CP",      f"CP_{m}_work"),
        ("트렌드포스",f"트렌드포스_{m}_work"),
        ("IDC",     f"IDC_{m}_work"),
        ("OmdiaTV", f"OmdiaTV_{m}_work"),
        ("DSCC",    f"DSCC_{m}_work"),
    ]

    # =========================
    # 1) 총평 표 수식 채우기
    # =========================
    start_row = 5
    for idx, (_label, sheet_name) in enumerate(sources):
        row = start_row + idx

        if sheet_name not in wb.sheetnames:
            continue

        summary_ws.cell(row=row, column=4).value = f"={sheet_name}!F2"  # D
        summary_ws.cell(row=row, column=5).value = f'=COUNTIF({sheet_name}!D5:D1048576,"연합뉴스")'  # E
        summary_ws.cell(row=row, column=6).value = f"={sheet_name}!F3"  # F
        summary_ws.cell(row=row, column=7).value = f"={sheet_name}!F4"  # G

    # =========================
    # 2) 차트 목록
    # =========================
    charts = getattr(summary_ws, "_charts", None) or []

    # =========================
    # 3) Chart1 데이터 범위 지정
    # =========================
    if len(charts) >= 1:
        try:
            chart1 = charts[0]

            categories = Reference(summary_ws, min_col=3, min_row=12, max_row=16)  # C12:C16
            data = Reference(summary_ws, min_col=4, min_row=11, max_col=5, max_row=16)  # D11:E16
        

            chart1.series = []
            chart1.add_data(data, titles_from_data=True)
            chart1.set_categories(categories)
        except Exception:
            pass

    # =========================
    # 4) 이번달 행 계산 + 값 링크 입력
    # =========================
    year, month = _seoul_now_year_month()
    target_row = _calc_year_month_row(year, month, base_year=2020, base_row=28)
    prev_row = target_row - 1

    # D~G: 이번달 값은 D5~D8 링크
    summary_ws[f"D{target_row}"].value = "=D5"  # CP
    summary_ws[f"E{target_row}"].value = "=D6"  # 트렌드포스
    summary_ws[f"F{target_row}"].value = "=D7"  # IDC
    summary_ws[f"G{target_row}"].value = "=D8"  # OmdiaTV
    summary_ws[f"H{target_row}"].value = "=D9"  # DSCC

    # =========================
    # 5) 증감 계산 (J24:N24) + 서식(빨강/bold)
    #    전월 셀이 비어있어도 에러 안 나게 IFERROR 처리
    # =========================
    red_bold = Font(color="FF0000", bold=True)

    summary_ws["J24"].value = f"=IFERROR(D{target_row}-D{prev_row}, D{target_row})"
    summary_ws["K24"].value = f"=IFERROR(E{target_row}-E{prev_row}, E{target_row})"
    summary_ws["L24"].value = f"=IFERROR(F{target_row}-F{prev_row}, F{target_row})"
    summary_ws["M24"].value = f"=IFERROR(G{target_row}-G{prev_row}, G{target_row})"
    summary_ws["N24"].value = f"=IFERROR(H{target_row}-H{prev_row}, H{target_row})"

    for addr in ("J24", "K24", "L24", "M24", "N24"):
        summary_ws[addr].font = red_bold

    for r in range(start_row, target_row + 1):
        c = summary_ws.cell(row=r, column=3)  # C열
        c.number_format = '0"월"'

    # =========================
    # 6) Chart2 데이터 범위 지정: 작년 1월 ~ 이번달
    # =========================
    if len(charts) >= 2:
        try:
            chart2 = charts[1]

            start_row = _calc_year_month_row(year - 1, 1, base_year=2020, base_row=29)

            # (선택) C열을 "n월"로 보이게 서식 지정
            for r in range(start_row, target_row + 1):
                summary_ws.cell(row=r, column=3).number_format = '0"월"'  # C열

            # categories: C(start)~C(target)
            cat_ref = Reference(
                summary_ws,
                min_col=3,  # C
                min_row=start_row,
                max_row=target_row,
            )

            # ✅ 중요: series 먼저 만들고, categories는 마지막에 설정
            chart2.series = []

            for col in range(4, 8):  # D(4)~G(7)
                values = Reference(
                    summary_ws,
                    min_col=col,
                    min_row=start_row,
                    max_row=target_row,
                )
                title = summary_ws.cell(row=24, column=col).value  # D24:G24 고정
                chart2.series.append(Series(values, title=title))

            # ✅ 마지막에 카테고리 세팅해야 모든 시리즈에 적용됨
            chart2.set_categories(cat_ref)

        except Exception:
            pass

def process_tracking_from_work(checked_bytes: bytes) -> bytes:
    wb = load_workbook(BytesIO(checked_bytes), data_only=False)

    for main_name, work_name in _find_main_and_work_pairs(wb):
        main_ws = wb[main_name]
        work_ws = wb[work_name]

        # 1. 원본 → _work 시트 기사 데이터 복사
        _copy_main_to_work(main_ws, work_ws)

        # 3. Tier Table 업데이트 (조건부)
        _update_tier_table_if_needed(wb, work_ws)

        # 4. 카테고리/건수 테이블 생성
        _fill_categories_and_counts(main_ws, work_ws)

        # 5. 건수 내림차순 정렬 (M/N)
        _sort_counts_to_MN(work_ws)

        # 6. 원그래프용 데이터(P/Q) 생성
        _prepare_chart_area(work_ws)

    _update_month_summary_sheet(wb)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()
