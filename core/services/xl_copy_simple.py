import calendar
from datetime import datetime
from io import BytesIO
import re
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# ===================== 분류 체계(범위 + 시장) =====================
GEOS = ["전세계", "한국", "중국", "유럽", "미국", "일본", "인도"]
MARKETS = [
    "스마트폰", "폴더블 스마트폰", "스마트폰 AP",
    "AI", "XR", "스마트워치",
    "보안",
    "디스플레이",
    "로봇청소기", "로봇",
    "반도체",
    "전기차",
]

ALLOWED_CATEGORIES = {f"{g} {m} 시장" for g in GEOS for m in MARKETS}
ALLOWED_CATEGORIES.add("미분류")


def _to_whitelist(cat: str) -> str:
    return cat if cat in ALLOWED_CATEGORIES else "미분류"


# ===================== 핵심: 시장 시그널 / 브랜드 시그널 =====================
MARKET_SIGNAL_RE = re.compile(
    r"(시장|점유율|출하량|판매량|판매|누적|ASP|평균판매단가|매출|성장률|CAGR|전년\s*동기|QoQ|YoY|전분기|분기|상반기|하반기)",
    re.I
)

PHONE_BRAND_RE = re.compile(r"(아이폰|iphone|갤럭시|galaxy|삼성전자|삼성|애플|apple)", re.I)
PHONE_SALES_RE = re.compile(r"(판매|판매량|출하|출하량|점유율|누적\s*판매|매출)", re.I)

AI_SMARTPHONE_RE = re.compile(r"(AI\s*스마트폰|생성형\s*AI\s*스마트폰|generative\s*AI\s*smartphone)", re.I)

# 디스플레이 분류 키워드
DISPLAY_FORCE_RE = re.compile(
    r"(TV|OLED|LCD|디스플레이|display|모니터|monitor|올레드|맥북|노트북)",
    re.I
)

# ===================== 범위(geo) 패턴 =====================
GEO_PATTERNS = {
    "전세계": r"(전\s*세계.{0,20}시장|글로벌.{0,20}시장|전\s*세계|전세계|세계|글로벌|global|worldwide)",
    "한국":   r"(한국.{0,20}시장|대한민국.{0,20}시장|국내.{0,20}시장|한국|대한민국|\b국내\b|Korea|South\s*Korea)",
    "중국":   r"(중국.{0,20}시장|중국|China)",
    "유럽":   r"(유럽.{0,20}시장|유럽|Europe|\bEU\b)",
    "미국":   r"(미국.{0,20}시장|미국|USA|\bUS\b|United\s*States|U\.S\.)",
    "일본":   r"(일본.{0,20}시장|일본|Japan|\bJP\b|도쿄|Tokyo)",
    "인도":   r"(인도.{0,20}시장|인도|India|델리|뭄바이|Mumbai|Bengaluru|Bangalore)",
}

# ===================== 시장(domain) 패턴 =====================
DOMAIN_PATTERNS = {
    "폴더블 스마트폰": r"(폴더블\s*스마트폰|폴더블|플립|플립폰|폴드|flip\b|fold\b|razr|레이저)",
    "스마트폰 AP": r"(\bAP\b|모바일\s*AP|\bSoC\b|chipset|칩셋|AP\s*원가|AP\s*비용|AP\s*공정)",

    "스마트폰": r"(스마트폰|smart\s*phone|smartphone|삼성폰|애플폰|mobile\s*phone|휴대폰)",

    "AI": r"(\bAI\b|인공지능|생성형\s*AI|Generative\s*AI|LLM|ChatGPT|Copilot|Gemini)",

    "XR": r"(\bXR\b|\bAR\b|\bVR\b|\bMR\b|헤드셋|스마트\s*안경|스마트안경)",
    "스마트워치": r"(스마트\s*워치|smart\s*watch|웨어러블|wearable)",

    "보안": r"(보안|사이버\s*보안|사이버\s*위협|cyber\s*security|security|침해|해킹)",

    "디스플레이": r"((디스플레이|display|모니터|monitor|게이밍\s*모니터)\s*시장)",

    "로봇청소기": r"(로봇\s*청소기|청소\s*로봇|robot\s*vacuum|vacuum\s*robot|로보락|Roborock|Ecovacs|Dreame)",
    "로봇": r"(로봇\b|로봇공학|서비스\s*로봇|산업용\s*로봇|제조용\s*로봇|로봇산업|휴머노이드|humanoid)",

    "반도체": r"(반도체|파운드리|foundry|칩\b|chips\b|chip\b|메모리|memory|HBM|\bDRAM\b|D-?RAM|\bNAND\b|"
             r"D램|디램|하이닉스|SK\s*하이닉스|엔비디아|NVIDIA|AMD|인텔|Intel|TSMC|마이크론|Micron|wafer|fab|패키징|CUDA)",
    "전기차": r"(전기차\b|전기차\s*시장|electric\s*vehicle|\bEV\b|\bBEV\b|\bPHEV\b)",
}

DOMAIN_PRIORITY = [
    "폴더블 스마트폰", "스마트폰 AP",
    "OLED", "LCD TV", "TV",
    "XR", "스마트워치",
    "보안",
    "로봇청소기", "로봇",
    "반도체",
    "전기차",
    "디스플레이",
    "AI",
    "스마트폰",
]


# ===================== 명시적 "<범위><시장> 시장" 최우선 탐지 =====================
def _compile_explicit_patterns():
    geo_tokens = GEO_PATTERNS.copy()
    market_tokens = {
        "스마트폰": r"(스마트폰|smart\s*phone|삼성폰|애플폰|휴대폰|mobile\s*phone)",
        "폴더블 스마트폰": r"(폴더블\s*스마트폰|폴더블|플립|폴드|flip\b|fold\b|razr|레이저)",
        "스마트폰 AP": r"(\bAP\b|모바일\s*AP|\bSoC\b|chipset|칩셋)",
        "AI": r"(\bAI\b|인공지능|생성형\s*AI|Generative\s*AI|LLM)",
        "XR": r"(\bXR\b|\bAR\b|\bVR\b|\bMR\b|헤드셋|스마트\s*안경|스마트안경)",
        "스마트워치": r"(스마트\s*워치|smart\s*watch|웨어러블|wearable)",
        "보안": r"(보안|사이버\s*보안|사이버\s*위협|cyber\s*security|security)",
        "디스플레이": r"(디스플레이|모니터)",
        "로봇청소기": r"(로봇\s*청소기|청소\s*로봇|robot\s*vacuum|Roborock|Ecovacs|Dreame)",
        "로봇": r"(로봇|휴머노이드|humanoid)",
        "반도체": r"(반도체|파운드리|foundry|메모리|HBM|DRAM|NAND|하이닉스|엔비디아|TSMC|Micron|CUDA)",
        "전기차": r"(전기차|electric\s*vehicle|\bEV\b|\bBEV\b|\bPHEV\b)",
    }

    patterns = []
    for g, gtok in geo_tokens.items():
        for m, mtok in market_tokens.items():
            p1 = rf"({gtok}).{{0,30}}({mtok}).{{0,10}}시장"
            p2 = rf"({mtok}).{{0,10}}시장.{{0,30}}({gtok})"
            patterns.append((g, m, re.compile(p1, re.I)))
            patterns.append((g, m, re.compile(p2, re.I)))
    return patterns


EXPLICIT_PATTERNS = _compile_explicit_patterns()


# ===================== 다중 국가 → '전세계' 강제 =====================
_GEO_TOKEN_SPECIFIC = [
    r"(한국|대한민국|\b국내\b|Korea|South\s*Korea)",
    r"(중국|China)",
    r"(유럽|Europe|\bEU\b)",
    r"(미국|USA|\bUS\b|United\s*States|U\.S\.)",
    r"(일본|Japan|\bJP\b|Tokyo|도쿄)",
    r"(인도|India|Mumbai|Bengaluru|Delhi|뭄바이|델리)",
]
_GEO_TOKEN_GLOBAL = r"(전\s*세계|전세계|글로벌|global|worldwide)"
GEO_RE_SPECIFICS = [re.compile(p, re.I) for p in _GEO_TOKEN_SPECIFIC]
GEO_RE_GLOBAL = re.compile(_GEO_TOKEN_GLOBAL, re.I)


def _multi_geo_triggers_world(text: str) -> bool:
    t = text or ""
    specific_hits = 0
    for rx in GEO_RE_SPECIFICS:
        if rx.search(t):
            specific_hits += 1
        if specific_hits >= 2:
            return True
    if specific_hits >= 1 and GEO_RE_GLOBAL.search(t):
        return True
    return False


def _find_explicit_geo_market(text: str):
    """
    본문에 '<범위><시장> 시장' 명시가 있으면 해당 조합 즉시 반환.
    단, 다중 국가 규칙이 트리거되면 범위는 '전세계'로 강제.
    """
    t = text or ""
    if _multi_geo_triggers_world(t):
        for g, m, rx in EXPLICIT_PATTERNS:
            if rx.search(t):
                return "전세계", m
        return "전세계", None

    for g, m, rx in EXPLICIT_PATTERNS:
        if rx.search(t):
            return g, m

    return None, None


# ===================== 공통 유틸 =====================
def _copy_range_values(src_ws, dst_ws, src_range: str, dst_top_left: str):
    min_col, min_row, max_col, max_row = range_boundaries(src_range)
    col_letters = "".join([c for c in dst_top_left if c.isalpha()])
    row_digits = "".join([c for c in dst_top_left if c.isdigit()])
    dst_row0 = int(row_digits)
    dst_col0 = 0
    for i, ch in enumerate(reversed(col_letters.upper())):
        dst_col0 += (ord(ch) - 64) * (26 ** i)
    rows = max_row - min_row + 1
    cols = max_col - min_col + 1
    for r in range(rows):
        for c in range(cols):
            val = src_ws.cell(row=min_row + r, column=min_col + c).value
            dst_ws.cell(row=dst_row0 + r, column=dst_col0 + c, value=val)


def _clear_range(ws, cell_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c, value=None)


def _find_last_row_by_column(ws, *, col: int, start_row: int, end_row: int) -> int:
    last_row = start_row - 1
    for r in range(start_row, end_row + 1):
        val = ws.cell(row=r, column=col).value
        if val is None or str(val).strip() == "":
            break
        last_row = r
    return last_row


def _rename_if_exists(wb, candidates, new_name):
    for name in candidates:
        if name in wb.sheetnames:
            wb[name].title = new_name
            return True
    last = candidates[-1] if candidates else None
    if last and last.startswith("re:"):
        rx = re.compile(last[3:], re.I)
        for s in wb.sheetnames:
            if rx.match(s):
                wb[s].title = new_name
                return True
    return False


def _fill_auto_numbers(ws, start_row: int = 5, col: int = 1, max_rows: int = 1500):
    count = 0
    for i in range(start_row, max_rows + 1):
        val = ws.cell(row=i, column=2).value
        if val is None or str(val).strip() == "":
            break
        count += 1
        ws.cell(row=i, column=col, value=count)


def _update_countif_formulas(ws, month, base_sheet="CP"):
    for row in range(7, 501):
        ws[f"K{row}"] = f'=COUNTIF({base_sheet}_{month}!G:G,L{row})'


# ===================== 크롤링(인터넷) =====================
def _fetch_article_text(url: str) -> str:
    if not url or not isinstance(url, str) or not url.startswith(("http://", "https://")):
        return ""
    try:
        res = requests.get(url.strip(), timeout=5, headers={"User-Agent": "Mozilla/5.0"})
        if res.status_code != 200:
            return ""
        soup = BeautifulSoup(res.text, "html.parser")

        for sel in ["header", "nav", "footer", "aside", "script", "style",
                    ".sidebar", ".breadcrumbs", ".breadcrumb", ".related",
                    ".recommend", ".ad", ".ads"]:
            for n in soup.select(sel):
                n.decompose()

        selectors = [
            "article", ".article", "#articleBody", "#articeBody", "#news_body",
            ".news_body", ".article_body", ".article-body", ".content", "#content",
            ".post-content", ".entry-content", ".post_body", ".post-body"
        ]
        nodes = []
        for sel in selectors:
            nodes = soup.select(sel)
            if nodes:
                break

        if nodes:
            text = " ".join(n.get_text(separator=" ", strip=True) for n in nodes)
        else:
            text = soup.get_text(separator=" ", strip=True)

        return (text or "")[:4000]
    except Exception:
        return ""


# ===================== 카테고리 분류 + reason =====================
def _regex_search(pattern, text):
    return re.search(pattern, text or "", flags=re.I) is not None


def _pick_geo(text):
    if _multi_geo_triggers_world(text):
        return "전세계"
    for geo, patt in GEO_PATTERNS.items():
        if _regex_search(patt, text):
            return geo
    return None


def _compose_category(geo_label, domain_label):
    if not domain_label:
        return "미분류"
    geo = geo_label if geo_label in GEOS else "전세계"
    return f"{geo} {domain_label} 시장"


def _pick_domains(text: str):
    hits = []
    for key in DOMAIN_PRIORITY:
        patt = DOMAIN_PATTERNS.get(key)
        if patt and _regex_search(patt, text):
            hits.append(key)
    return hits


def _is_market_article(text: str) -> bool:
    return MARKET_SIGNAL_RE.search(text or "") is not None


def _resolve_representative_domain(text: str, domains: list[str]) -> tuple[str | None, str]:
    """
    대표 도메인 결정 + reason
    """
    if not domains:
        if PHONE_BRAND_RE.search(text or "") and PHONE_SALES_RE.search(text or ""):
            return "스마트폰", "R_BRAND_SALES_TO_SMARTPHONE"
        return None, "R_NO_DOMAIN_MATCH"

    s = set(domains)

    if "AI" in s and ("스마트폰" in s or AI_SMARTPHONE_RE.search(text or "")):
        if "폴더블 스마트폰" in s:
            return "폴더블 스마트폰", "R_AI_SMARTPHONE_TO_FOLDABLE"
        return "스마트폰", "R_AI_SMARTPHONE_TO_SMARTPHONE"

    if "폴더블 스마트폰" in s:
        return "폴더블 스마트폰", "R_DOMAIN_PRIORITY_FOLDABLE"
    if "스마트폰 AP" in s:
        return "스마트폰 AP", "R_DOMAIN_PRIORITY_PHONE_AP"
    if "로봇청소기" in s:
        return "로봇청소기", "R_DOMAIN_PRIORITY_ROBOT_VAC"
    if "반도체" in s:
        return "반도체", "R_DOMAIN_PRIORITY_SEMI"
    if "디스플레이" in s and _is_market_article(text):
        return "디스플레이", "R_DOMAIN_DISPLAY_WITH_SIGNAL"
    if "디스플레이" in s and not _is_market_article(text):
        s.remove("디스플레이")
    if "AI" in s and not _is_market_article(text):
        s.remove("AI")
    if "스마트폰" in s and _is_market_article(text):
        return "스마트폰", "R_SMARTPHONE_WITH_SIGNAL"
    for key in ["보안", "XR", "스마트워치", "로봇", "전기차"]:
        if key in s and _is_market_article(text):
            return key, f"R_{key}_WITH_SIGNAL"
        if key in s and not _is_market_article(text):
            s.remove(key)
    for key in DOMAIN_PRIORITY:
        if key in s:
            if key == "스마트폰" and not _is_market_article(text):
                if PHONE_BRAND_RE.search(text or "") and PHONE_SALES_RE.search(text or ""):
                    return "스마트폰", "R_PHONE_BRAND_SALES_FALLBACK"
                return None, "R_SMARTPHONE_NO_SIGNAL"
            return key, "R_DOMAIN_FALLBACK_PRIORITY"

    return None, "R_DOMAIN_EMPTY_AFTER_FILTER"


def classify_with_reason(text: str, source_hint: str) -> tuple[str, str]:
    t = text or ""

    # =========================
    # - TV/OLED/LCD/디스플레이/모니터/올레드 단어가 포함되면 무조건 "디스플레이 시장"
    # - 국가/지역 언급 없거나 2개 이상이면 "전세계 디스플레이 시장"
    # - 1개 국가면 "{국가} 디스플레이 시장"
    # =========================
    if DISPLAY_FORCE_RE.search(t):
        geo = _pick_geo(t) or "전세계"
        cat = _to_whitelist(_compose_category(geo, "디스플레이"))
        return cat, "R_FORCE_DISPLAY"

    # (1) 명시 "<범위><시장> 시장" 최우선 (디스플레이 강제 규칙 다음)
    eg, em = _find_explicit_geo_market(t)
    if em:
        cat = _to_whitelist(_compose_category(eg if eg else "전세계", em))
        return cat, "R_EXPLICIT_GEO_MARKET"

    # (2) geo
    geo = _pick_geo(t) or "전세계"

    # (3) 도메인 후보 수집
    domains = _pick_domains(t)

    if _is_market_article(t) and _regex_search(DOMAIN_PATTERNS["스마트폰"], t):
        if "스마트폰" not in domains:
            domains.append("스마트폰")

    for k in ["반도체", "전기차", "로봇", "로봇청소기"]:
        if _is_market_article(t) and _regex_search(DOMAIN_PATTERNS[k], t) and k not in domains:
            domains.append(k)

    rep_domain, reason = _resolve_representative_domain(t, domains)

    if not rep_domain and source_hint in ("OmdiaTV", "DSCC"):
        return _to_whitelist(_compose_category(geo, "디스플레이")), "R_SOURCE_HINT_DISPLAY"

    if rep_domain:
        return _to_whitelist(_compose_category(geo, rep_domain)), reason

    return "미분류", reason


# ===================== 카테고리 입력 루틴 (G=카테고리, H=reason) =====================
def _fill_categories(ws, source_hint: str, start_row: int = 5, max_rows: int | None = None):
    """
    max_rows가 None이면 시트의 실제 max_row까지 자동 처리.
    (기존처럼 800/2500 같은 상한 때문에 누락되는 문제 제거)
    """
    if max_rows is None:
        max_rows = ws.max_row or start_row

    for r in range(start_row, max_rows + 1):
        bval = ws.cell(row=r, column=2).value
        e_text = ws.cell(row=r, column=5).value
        f_url = ws.cell(row=r, column=6).value
        if not (bval or e_text or f_url):
            break

        text_source = (str(e_text).strip() if e_text else "")

        if len(text_source) < 100 and f_url:
            fetched = _fetch_article_text(str(f_url))
            if fetched:
                text_source = fetched

        cat, reason = classify_with_reason(text_source, source_hint)

        ws.cell(row=r, column=7, value=cat)     # G열
        ws.cell(row=r, column=8, value=reason)  # H열


# ===================== 테스트/시뮬레이터(모의 테스트) =====================
def simulate_classification(text_list: list[str], source_hint: str = "CP") -> list[dict]:
    out = []
    for t in text_list:
        cat, reason = classify_with_reason(t, source_hint)
        out.append({"text": t[:120], "category": cat, "reason": reason})
    return out


# ===================== 메인 처리 =====================
def process_monthly_copy(raw_bytes: bytes, monthly_bytes: bytes, month: int) -> bytes:
    raw_wb = load_workbook(BytesIO(raw_bytes), data_only=True)
    mon_wb = load_workbook(BytesIO(monthly_bytes))
    m = int(month)

    # 시트명 변경
    _rename_if_exists(mon_wb, ["CP_9", "cp_9", "CP-9", "re:^CP[_ -]?9$"], f"CP_{m}")
    _rename_if_exists(mon_wb, ["CP_9_work", "CP_9_Work", "re:^CP[_ -]?9[_ -]?work$"], f"CP_{m}_work")
    _rename_if_exists(mon_wb, ["트렌드포스_9", "트렌드포스 9", "re:^트렌드포스[_ -]?9$"], f"트렌드포스_{m}")
    _rename_if_exists(mon_wb, ["트렌드포스_9_work", "트렌드포스_9_Work", "re:^트렌드포스[_ -]?9[_ -]?work$"], f"트렌드포스_{m}_work")
    _rename_if_exists(mon_wb, ["IDC_9", "re:^IDC[_ -]?9$"], f"IDC_{m}")
    _rename_if_exists(mon_wb, ["IDC_9_work", "re:^IDC[_ -]?9[_ -]?work$"], f"IDC_{m}_work")
    _rename_if_exists(mon_wb, ["OmdiaTV_9", "Omdia TV_9", "re:^Omdia\s?TV[_ -]?9$"], f"OmdiaTV_{m}")
    _rename_if_exists(mon_wb, ["OmdiaTV_9_work", "Omdia TV_9_work", "re:^Omdia\s?TV[_ -]?9[_ -]?work$"], f"OmdiaTV_{m}_work")
    _rename_if_exists(mon_wb, ["DSCC_9", "re:^DSCC[_ -]?9$"], f"DSCC_{m}")
    _rename_if_exists(mon_wb, ["DSCC_9_work", "re:^DSCC[_ -]?9[_ -]?work$"], f"DSCC_{m}_work")
    _rename_if_exists(mon_wb, [f"9월 총평", "re:^9\s*월\s*총평$"], f"{m}월 총평")

    sources = [
        ("CPR", "CP"),
        ("트렌드포스", "트렌드포스"),
        ("IDC", "IDC"),
        ("Omdia TV", "OmdiaTV"),
        ("DSCC", "DSCC"),
    ]

    for raw_name, target_prefix in sources:
        # RAW에 해당 시트가 없으면 스킵
        if raw_name not in raw_wb.sheetnames:
            continue

        src_ws = raw_wb[raw_name]
        dst_name = f"{target_prefix}_{m}"
        if dst_name not in mon_wb.sheetnames:
            continue
        dst_ws = mon_wb[dst_name]

        _clear_range(dst_ws, "B5:B2000")
        _clear_range(dst_ws, "C5:F2000")

        last_row = _find_last_row_by_column(src_ws, col=2, start_row=5, end_row=2000)
        if last_row < 5:
            continue

        _copy_range_values(src_ws, dst_ws, f"B5:B{last_row}", "B5")
        _copy_range_values(src_ws, dst_ws, f"D5:G{last_row}", "C5")

    # 번호 매기기
    for name in [f"CP_{m}", f"트렌드포스_{m}", f"IDC_{m}", f"OmdiaTV_{m}", f"DSCC_{m}"]:
        if name in mon_wb.sheetnames:
            _fill_auto_numbers(mon_wb[name], max_rows=3000)

    # A2 수식 업데이트
    for name in [f"트렌드포스_{m}", f"IDC_{m}", f"OmdiaTV_{m}", f"DSCC_{m}"]:
        if name in mon_wb.sheetnames:
            mon_wb[name]["A2"] = f"=CP_{m}!A2"

    # ✅ CP_{m} 시트 A2 날짜 자동 업데이트
    current_year = datetime.now().year
    last_day = calendar.monthrange(current_year, m)[1]
    start_date = f"{current_year}/{m:02d}/01"
    end_date = f"{current_year}/{m:02d}/{last_day:02d}"
    if f"CP_{m}" in mon_wb.sheetnames:
        mon_wb[f"CP_{m}"]["A2"] = f"[기간] {start_date}~ {end_date}"

    # 카테고리 자동 분류 (G열), reason(H열)
    if f"CP_{m}" in mon_wb.sheetnames:
        _fill_categories(mon_wb[f"CP_{m}"], "CP") 
    if f"트렌드포스_{m}" in mon_wb.sheetnames:
        _fill_categories(mon_wb[f"트렌드포스_{m}"], "트렌드포스")
    if f"IDC_{m}" in mon_wb.sheetnames:
        _fill_categories(mon_wb[f"IDC_{m}"], "IDC")
    if f"OmdiaTV_{m}" in mon_wb.sheetnames:
        _fill_categories(mon_wb[f"OmdiaTV_{m}"], "OmdiaTV")
    if f"DSCC_{m}" in mon_wb.sheetnames:
        _fill_categories(mon_wb[f"DSCC_{m}"], "DSCC")

    out = BytesIO()
    mon_wb.save(out)
    out.seek(0)
    return out.getvalue()